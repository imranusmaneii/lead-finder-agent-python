"""Tests for build_excel module."""

import sys
import os
import tempfile

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from build_excel import build_excel, get_excel_filename, _sanitize_filename
from openpyxl import load_workbook


class TestSanitizeFilename:
    def test_basic(self):
        assert _sanitize_filename("Coffee Shops") == "coffee_shops"

    def test_special_chars(self):
        assert _sanitize_filename("cafe & bar!") == "cafe_bar"

    def test_empty(self):
        assert _sanitize_filename("") == ""

    def test_multiple_spaces(self):
        assert _sanitize_filename("burger  shop") == "burger_shop"


class TestGetExcelFilename:
    def test_both_category_and_location(self):
        assert get_excel_filename("coffee shops", "America") == "leads_coffee_shops_america.xlsx"

    def test_category_only(self):
        assert get_excel_filename("restaurants", "") == "leads_restaurants.xlsx"

    def test_location_only(self):
        assert get_excel_filename("", "New York") == "leads_new_york.xlsx"

    def test_neither(self):
        assert get_excel_filename("", "") == "leads.xlsx"


class TestBuildExcel:
    def _sample_leads(self):
        return [
            {
                "business_name": "Coffee House",
                "email": "info@coffeehouse.com",
                "phone": "+1-555-1234",
                "website": "https://coffeehouse.com",
                "location": "123 Main St",
            },
            {
                "business_name": "Bean Cafe",
                "email": "",
                "phone": "+1-555-5678",
                "website": "https://beancafe.com",
                "location": "456 Oak Ave",
            },
        ]

    def test_creates_file(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            original_dir = os.getcwd()
            os.chdir(tmpdir)
            try:
                filename = build_excel(self._sample_leads(), "coffee shops", "America")
                assert os.path.exists(filename)
                assert filename == "leads_coffee_shops_america.xlsx"
            finally:
                os.chdir(original_dir)

    def test_headers_are_correct(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            original_dir = os.getcwd()
            os.chdir(tmpdir)
            try:
                filename = build_excel(self._sample_leads(), "coffee", "NYC")
                wb = load_workbook(filename)
                ws = wb.active
                headers = [ws.cell(row=1, column=c).value for c in range(1, 6)]
                assert headers == ["Business Name", "Email", "Phone Number", "Website", "Location"]
            finally:
                os.chdir(original_dir)

    def test_data_rows(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            original_dir = os.getcwd()
            os.chdir(tmpdir)
            try:
                filename = build_excel(self._sample_leads(), "coffee", "NYC")
                wb = load_workbook(filename)
                ws = wb.active
                # Row 2
                assert ws.cell(row=2, column=1).value == "Coffee House"
                assert ws.cell(row=2, column=2).value == "info@coffeehouse.com"
                assert ws.cell(row=2, column=3).value == "+1-555-1234"
                # Row 3
                assert ws.cell(row=3, column=1).value == "Bean Cafe"
                assert ws.cell(row=3, column=2).value in ("", None)
            finally:
                os.chdir(original_dir)

    def test_empty_leads(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            original_dir = os.getcwd()
            os.chdir(tmpdir)
            try:
                filename = build_excel([], "coffee", "NYC")
                wb = load_workbook(filename)
                ws = wb.active
                # Only header row
                assert ws.cell(row=1, column=1).value == "Business Name"
                assert ws.cell(row=2, column=1).value is None
            finally:
                os.chdir(original_dir)

    def test_header_styling(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            original_dir = os.getcwd()
            os.chdir(tmpdir)
            try:
                filename = build_excel(self._sample_leads(), "coffee", "NYC")
                wb = load_workbook(filename)
                ws = wb.active
                cell = ws.cell(row=1, column=1)
                assert cell.font.bold is True
                assert cell.font.color.rgb == "00FFFFFF"
                assert cell.fill.start_color.rgb == "004472C4"
            finally:
                os.chdir(original_dir)

"""Lead Finder Agent — CLI entry point.

Usage:
    python main.py
    python main.py "coffee shops in America"
    python main.py --prompt "dentists in New York" --output results.xlsx

The agent:
1. Parses the user prompt into business category + location
2. Scrapes Bing Maps for matching businesses
3. Falls back to Google Search if Bing Maps returns no results
4. Extracts business name, phone, website, address, and email
5. Saves results to an Excel file
6. Prints a summary
"""

import argparse
import sys
from parse_prompt import parse_prompt
from scrape_leads import scrape_leads
from build_excel import build_excel, get_excel_filename

__version__ = "1.0.0"


def main() -> None:
    parser = argparse.ArgumentParser(
        prog="lead-finder",
        description="Find business leads from Bing Maps / Google Search.",
    )
    parser.add_argument(
        "prompt",
        nargs="*",
        help='Search prompt, e.g. "coffee shops in America"',
    )
    parser.add_argument(
        "-p", "--prompt-flag",
        dest="prompt_flag",
        help="Alternative way to pass the search prompt",
    )
    parser.add_argument(
        "-o", "--output",
        help="Custom output Excel filename (default: auto-generated)",
    )
    parser.add_argument(
        "-v", "--version",
        action="version",
        version=f"%(prog)s {__version__}",
    )
    args = parser.parse_args()

    # Determine the prompt from positional args or --prompt-flag
    if args.prompt_flag:
        prompt = args.prompt_flag
    elif args.prompt:
        prompt = " ".join(args.prompt)
    else:
        prompt = input("Enter your search prompt (e.g. 'coffee shops in America'): ").strip()

    if not prompt:
        print("Error: No prompt provided.")
        sys.exit(1)

    print(f"\n{'='*60}")
    print(f"  Lead Finder Agent v{__version__}")
    print(f"{'='*60}")
    print(f"  Prompt: {prompt}")

    # Step 1: Parse the prompt
    business_category, location = parse_prompt(prompt)
    if not business_category:
        print("Error: Could not understand the search prompt.")
        print("Try something like 'coffee shops in New York' or 'dentists in Chicago'.")
        sys.exit(1)

    query_display = f"{business_category} in {location}" if location else business_category
    print(f"  Category: {business_category}")
    print(f"  Location: {location or '(not specified)'}")
    print(f"  Search query: {query_display}")
    print(f"{'='*60}\n")

    # Step 2: Scrape leads
    try:
        leads = scrape_leads(business_category, location)
    except Exception as e:
        print(f"\nError during scraping: {e}")
        print("Make sure Playwright and Chromium are installed:")
        print("  pip install -r requirements.txt")
        print("  playwright install chromium")
        sys.exit(1)

    # Step 3: Save to Excel
    if leads:
        if args.output:
            filename = args.output
            if not filename.endswith(".xlsx"):
                filename += ".xlsx"
            from openpyxl import Workbook
            from build_excel import COLUMNS, LEAD_KEYS
            wb = Workbook()
            ws = wb.active
            ws.title = "Leads"
            from openpyxl.styles import Font, PatternFill, Alignment
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            for col_idx, header in enumerate(COLUMNS, start=1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="left")
            for row_idx, lead in enumerate(leads, start=2):
                for col_idx, key in enumerate(LEAD_KEYS, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=lead.get(key, ""))
            col_widths = [30, 30, 20, 35, 40]
            for col_idx, width in enumerate(col_widths, start=1):
                ws.column_dimensions[chr(64 + col_idx)].width = width
            wb.save(filename)
        else:
            filename = build_excel(leads, business_category, location)
        print(f"\n  Excel file saved: {filename}")
    else:
        filename = get_excel_filename(business_category, location)
        print("\n  No leads found. No Excel file generated.")

    # Step 4: Print summary
    print(f"\n{'='*60}")
    print(f"  SUMMARY")
    print(f"{'='*60}")
    print(f"  Search query:        {query_display}")
    print(f"  Business category:   {business_category}")
    print(f"  Location:            {location or '(not specified)'}")
    print(f"  Leads collected:     {len(leads)}")
    if leads:
        print(f"  Excel file:          {filename}")
        print(f"  Leads with email:    {sum(1 for l in leads if l.get('email'))}")
        print(f"  Leads with phone:    {sum(1 for l in leads if l.get('phone'))}")
        print(f"  Leads with website:  {sum(1 for l in leads if l.get('website'))}")
    print(f"{'='*60}")


if __name__ == "__main__":
    main()

"""Build an Excel (.xlsx) file from a list of leads using openpyxl."""

import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


COLUMNS = ["Business Name", "Email", "Phone Number", "Website", "Location"]
LEAD_KEYS = ["business_name", "email", "phone", "website", "location"]


def _sanitize_filename(text: str) -> str:
    """Convert text to a safe filename segment."""
    return re.sub(r"[^a-z0-9]+", "_", text.lower()).strip("_")


def get_excel_filename(category: str, location: str) -> str:
    """Generate a meaningful filename for the Excel output."""
    cat = _sanitize_filename(category)
    loc = _sanitize_filename(location)
    if not cat and not loc:
        return "leads.xlsx"
    if not loc:
        return f"leads_{cat}.xlsx"
    return f"leads_{cat}_{loc}.xlsx"


def build_excel(leads: list[dict], category: str, location: str) -> str:
    """Create an Excel file from leads and return the filename.

    The file is saved to the current working directory.
    Returns the filename used.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"

    # Write headers
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for col_idx, header in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left")

    # Write data rows
    for row_idx, lead in enumerate(leads, start=2):
        for col_idx, key in enumerate(LEAD_KEYS, start=1):
            ws.cell(row=row_idx, column=col_idx, value=lead.get(key, ""))

    # Auto-fit column widths (approximate)
    col_widths = [30, 30, 20, 35, 40]
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[chr(64 + col_idx)].width = width

    filename = get_excel_filename(category, location)
    wb.save(filename)
    return filename
